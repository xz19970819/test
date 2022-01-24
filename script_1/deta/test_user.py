'''
@File:test_user.py
@DateTime:2022/1/21 14:52 
@Author:eason 
@Desc:
'''

import openpyxl
from script_1.config.config import file,sheet
import os

class Table_user:
    def __init__(self,file,sheet):
        self.file = file
        self.sheet = sheet
        self.wb = openpyxl.load_workbook(self.file)
        self.table = self.wb[self.sheet]
        self.table = self.wb.active
        self.nrow = self.table.max_row
        self.colu = self.table.max_column
    def read(self):
        list2=[]
        for i in range(2,self.nrow+1):
            list1=[]
            for l in range(1,self.colu+1):
                value=self.table.cell(i,l).value
                list1.append(value)
            list2.append(list1)
        self.wb.close()
        return list2

    def write(self,*arg):
        for i in range(1,len(arg)+1):
            self.table.cell(self.nrow+1,i).value=arg[i-1]
        self.wb.save(self.file)
        self.wb.close()

if __name__ == '__main__':
    doc1=Table_user(file,sheet)
    print(doc1.read())
    # doc1.write("第一行","第二行")